# -*- coding: utf-8 -*-
# @Author   :hzj
# @File     :excel_demo.py
# @Time     :2020/6/29 15:56
import openpyxl
from Common.Driver import Driver
from openpyxl.styles import PatternFill, Font

# 读取excel文件
excel = openpyxl.load_workbook('../DataFile/WEBUI_demo.xlsx')
# 获取到需要的sheet页
sheetnames = excel.sheetnames
for sheetname in sheetnames:
    sheet = excel[sheetname]
    for value in sheet.values:
        if type(value[0]) is int:
            if value[1] == 'open_browser':
                driver = Driver(value[4])
            elif 'assert' in value[1]:
                status = getattr(driver, value[1])(value[2], value[3], value[6])
                if status is True:
                    # 写入pass，在第九行第八列写入
                    row = value[0] + 1
                    sheet.cell(row=row, column=8).value = 'Pass'
                    sheet.cell(row=row, column=8).fill = PatternFill('solid', fgColor='00FF00')
                    sheet.cell(row=row, column=8).font = Font(bold=True)
                else:
                    # 写入false
                    row = value[0] + 1
                    sheet.cell(row=row, column=8).value = 'False'
                    sheet.cell(row=row, column=8).fill = PatternFill('solid', fgColor='FF0000')
                    sheet.cell(row=row, column=8).font = Font(bold=True)
                excel.save('../DataFile/WEBUI_demo.xlsx')
            else:
                getattr(driver, value[1])(value[2], value[3], value[4])
        else:
            pass
'''
    excel文件驱动测试 v1.0
'''
# # 读取所有的excel内容
# for value in sheet.values:
#     if value[1] == 'open_browser':
#         driver = Driver(value[4])
#     elif value[1] == 'visit':
#         driver.visit(value[4])
#     elif value[1] == 'input':
#         driver.input(value[2], value[3], value[4])
#     elif value[1] == "click":
#         driver.click(value[2], value[3])
#     elif value[1] == 'wait_y':
#         driver.wait_y(value[4])
#     elif value[1] == 'assert_text':
#         status = driver.assert_text(value[2], value[3], value[6])
#         if status is True:
#             # 写入pass，在第九行第八列写入
#             row = value[0] + 1
#             sheet.cell(row=row, column=8).value = 'Pass'
#             sheet.cell(row=row, column=8).fill = PatternFill('solid', fgColor='00FF00')
#             sheet.cell(row=row, column=8).font = Font(bold=True)
#         else:
#             # 写入false
#             row = value[0] + 1
#             sheet.cell(row=row, column=8).value = 'False'
#             sheet.cell(row=row, column=8).fill = PatternFill('solid', fgColor='FF0000')
#             sheet.cell(row=row, column=8).font = Font(bold=True)
#         excel.save('../DataFile/WEBUI_demo.xlsx')
#     else:
#         pass
'''
    excel文件驱动测试 v2.0
'''