'''
    封装excel文件驱动UI自动化测试的类
'''
from Common.Driver import Driver
from Common.Log import Log
from openpyxl.styles import PatternFill, Font
import openpyxl

log = Log().logger


class ExcelDriver:
    index = 0

    def excel_driver(self, filepath):
        ec = ExcelOperator()
        excel = ec.excel_read(filepath)
        # 获取所有的sheet页
        sheets = excel.sheetnames
        try:
            for i in sheets:
                sheet = excel[i]
                ExcelDriver.index += 1
                log.info('执行UI测试第<{0}>个用例：{1}'.format(ExcelDriver.index, i))
                for value in sheet.values:
                    # 第一列的内容是int类型，即编号
                    if type(value[0]) is int:
                        log.info('执行关键字：{0}，操作描述：{1}'.format(value[1], value[5]))
                        # 第二列中的open_browser流程，走特殊的方法实例化driver对象
                        if value[1] == 'open_browser':
                            # 实例化driver
                            driver = Driver(value[4])
                        # 第二列存在assert这个字符串时，代表断言，需要进行判断传参
                        elif 'assert' in value[1]:
                            # 获取方法的参数个数，把self也计算在内
                            fuc = getattr(driver, value[1])
                            num = fuc.__code__.co_argcount
                            if num == 2:
                                status = getattr(driver, value[1])(value[6])
                            elif num == 4:
                                status = getattr(driver, value[1])(value[2], value[3], value[6])
                            else:
                                # 断言方法匹配不失败，则断言结果为False，避免status为空
                                status = False
                            # 获取断言结果，对结果进行判断
                            row = value[0] + 1
                            if status is True:
                                ec.cell_set(sheet, row, 'pass')
                            else:
                                ec.cell_set(sheet, row, 'false')
                            # excel文件保存机制
                            ec.excel_save(excel, filepath)
                        else:  # 除open_browser和断言外的其他流程，根据每列是否存在内容进行传参
                            # 获取方法的参数个数，把self也计算在内
                            fuc = getattr(driver, value[1])
                            num = fuc.__code__.co_argcount
                            if num == 1:
                                getattr(driver, value[1])()
                            elif num == 2:
                                getattr(driver, value[1])(value[4])
                            elif num == 3:
                                getattr(driver, value[1])(value[2], value[3])
                            elif num == 4:
                                getattr(driver, value[1])(value[2], value[3], value[4])
                            else:
                                log.error('方法配对失败！')

                    else:
                        pass
        except Exception as error:
            log.error('运行出现异常，异常信息描述为{}'.format(error))
        finally:
            # 不论执行清空如何，都要关闭文件
            log.info('文件读取完毕，自动化执行结束！\n')
            ec.excel_close(excel)


class ExcelOperator:
    '''
        Excel文件单元格设置，将断言一行的结果按实际情况进行设置
    '''
    def excel_read(self, filename):
        # 读取excel文件
        excel = openpyxl.load_workbook(filename)
        log.info('读取excel文件：{}'.format(filename))
        return excel

    def excel_close(self, excel):
        excel.close()

    def excel_save(self, excel, path):
        excel.save(path)

    def cell_set(self, sheet, row, value):
        bold = Font(bold=True)
        if value == 'pass':
            fill = PatternFill('solid', fgColor='AACF91')
        elif value == 'false':
            fill = PatternFill('solid', fgColor='FF0000')
        else:
            pass
        sheet.cell(row=row, column=8).value = value.upper()
        sheet.cell(row=row, column=8).fill = fill
        sheet.cell(row=row, column=8).font = bold
