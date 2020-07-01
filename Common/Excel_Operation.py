# -*- coding: utf-8 -*-
# @Author   :hzj
# @File     :Excel_Operation.py
# @Time     :2020/7/1 22:23
import openpyxl
from Common.Log import Log
from openpyxl.styles import PatternFill, Font

log = Log().logger


class ExcelOperator:
    def excel_read(self, filename):
        # 读取excel文件
        excel = openpyxl.load_workbook(filename)
        log.info('读取excel文件：{}'.format(filename))
        return excel

    def excel_close(self, excel):
        excel.close()

    def excel_save(self, excel, path):
        excel.save(path)

    def cell_set(self, sheet, row, value):
        bold = Font(bold=True)
        if value == 'pass':
            fill = PatternFill('solid', fgColor='AACF91')
        elif value == 'false':
            fill = PatternFill('solid', fgColor='FF0000')
        else:
            pass
        sheet.cell(row=row, column=8).value = value.upper()
        sheet.cell(row=row, column=8).fill = fill
        sheet.cell(row=row, column=8).font = bold
